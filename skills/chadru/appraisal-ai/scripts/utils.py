"""
Shared utilities for the appraisal drafting tool.
- Word tracked-change XML helpers (w:del + w:ins)
- Structured Markdown data loading/saving
"""

import copy
import datetime
import os
import re
import shutil
import tempfile
from lxml import etree


# ── Word XML namespaces ──────────────────────────────────────────────────────

NAMESPACES = {
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
    'w14': 'http://schemas.microsoft.com/office/word/2010/wordml',
}

# Register namespaces so lxml preserves prefixes on serialization
for prefix, uri in NAMESPACES.items():
    etree.register_namespace(prefix, uri)

W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'


# ── Tracked-change ID counter ────────────────────────────────────────────────

_change_id_counter = 0


def next_change_id():
    """Return an incrementing change ID for tracked-change markup."""
    global _change_id_counter
    _change_id_counter += 1
    return str(_change_id_counter)


def reset_change_id(start=0):
    """Reset the change ID counter (call at start of each document build)."""
    global _change_id_counter
    _change_id_counter = start


def get_author():
    return "Appraisal Tool"


def get_date():
    return datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")


# ── Tracked-change XML builders ──────────────────────────────────────────────

def make_run_with_rpr(text, rpr_element=None):
    """
    Create a <w:r> element containing <w:t> with the given text.
    If rpr_element is provided, clone it into the run to preserve formatting.
    """
    r = etree.SubElement(etree.Element('dummy'), f'{W}r')
    r.getparent().remove(r)
    r = etree.Element(f'{W}r')
    if rpr_element is not None:
        r.append(copy.deepcopy(rpr_element))
    t = etree.SubElement(r, f'{W}t')
    t.text = text
    # Preserve leading/trailing spaces
    if text and (text[0] == ' ' or text[-1] == ' '):
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    return r


def make_tracked_delete(run_element):
    """
    Wrap a <w:r> inside a <w:del> tracked-change element.
    Returns the <w:del> element.
    """
    del_elem = etree.Element(f'{W}del')
    del_elem.set(f'{W}id', next_change_id())
    del_elem.set(f'{W}author', get_author())
    del_elem.set(f'{W}date', get_date())

    # Convert <w:t> children to <w:delText>
    r_copy = copy.deepcopy(run_element)
    for t in r_copy.findall(f'.//{W}t'):
        t.tag = f'{W}delText'

    del_elem.append(r_copy)
    return del_elem


def make_tracked_insert(run_element):
    """
    Wrap a <w:r> inside a <w:ins> tracked-change element.
    Returns the <w:ins> element.
    """
    ins_elem = etree.Element(f'{W}ins')
    ins_elem.set(f'{W}id', next_change_id())
    ins_elem.set(f'{W}author', get_author())
    ins_elem.set(f'{W}date', get_date())
    ins_elem.append(copy.deepcopy(run_element))
    return ins_elem


def tracked_replace_in_run(run, old_text, new_text, parent):
    """
    Replace `old_text` with `new_text` inside a single <w:r> element using
    tracked changes (w:del old, w:ins new).

    `parent` is the paragraph (<w:p>) or other container holding the run.
    Modifies the parent in-place. Returns True if a replacement was made.
    """
    t_elem = run.find(f'{W}t')
    if t_elem is None or t_elem.text is None:
        return False

    if old_text not in t_elem.text:
        return False

    rpr = run.find(f'{W}rPr')
    full_text = t_elem.text
    idx = full_text.index(old_text)
    before = full_text[:idx]
    after = full_text[idx + len(old_text):]

    # Build replacement elements
    run_index = list(parent).index(run)
    parent.remove(run)

    insert_pos = run_index
    if before:
        r_before = make_run_with_rpr(before, rpr)
        parent.insert(insert_pos, r_before)
        insert_pos += 1

    # Tracked delete of old text
    r_old = make_run_with_rpr(old_text, rpr)
    del_elem = make_tracked_delete(r_old)
    parent.insert(insert_pos, del_elem)
    insert_pos += 1

    # Tracked insert of new text
    r_new = make_run_with_rpr(new_text, rpr)
    ins_elem = make_tracked_insert(r_new)
    parent.insert(insert_pos, ins_elem)
    insert_pos += 1

    if after:
        r_after = make_run_with_rpr(after, rpr)
        parent.insert(insert_pos, r_after)

    return True


def get_paragraph_text(paragraph):
    """Extract plain text from direct child <w:r> elements of a <w:p>.

    Skips runs nested inside <w:hyperlink> or other wrapper elements so that
    hyperlink text is handled separately by _replace_in_hyperlinks().
    """
    texts = []
    for child in paragraph:
        if child.tag == f'{W}r':
            for t in child.findall(f'{W}t'):
                if t.text:
                    texts.append(t.text)
    return ''.join(texts)


def _rpr_signature(run):
    """Return a hashable signature for a run's formatting properties."""
    rpr = run.find(f'{W}rPr')
    if rpr is None:
        return b''
    return etree.tostring(rpr, method='c14n2')


def merge_runs(paragraph):
    """
    Merge adjacent <w:r> elements that have identical formatting (rPr).
    This consolidates text that Word fragmented across multiple runs,
    making string matching reliable.
    """
    children = list(paragraph)
    i = 0
    while i < len(children) - 1:
        curr = children[i]
        nxt = children[i + 1]

        # Only merge if both are <w:r> elements
        if curr.tag != f'{W}r' or nxt.tag != f'{W}r':
            i += 1
            continue

        # Skip runs that contain non-text children (e.g. images, breaks)
        curr_has_special = any(
            child.tag not in (f'{W}t', f'{W}rPr')
            for child in curr
        )
        nxt_has_special = any(
            child.tag not in (f'{W}t', f'{W}rPr')
            for child in nxt
        )
        if curr_has_special or nxt_has_special:
            i += 1
            continue

        # Check if formatting matches
        if _rpr_signature(curr) != _rpr_signature(nxt):
            i += 1
            continue

        # Merge: append next run's text to current run's text
        curr_t = curr.find(f'{W}t')
        nxt_t = nxt.find(f'{W}t')

        curr_text = curr_t.text if curr_t is not None and curr_t.text else ''
        nxt_text = nxt_t.text if nxt_t is not None and nxt_t.text else ''

        merged_text = curr_text + nxt_text

        if curr_t is None:
            curr_t = etree.SubElement(curr, f'{W}t')
        curr_t.text = merged_text

        # Preserve spaces
        if merged_text and (merged_text[0] == ' ' or merged_text[-1] == ' '):
            curr_t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

        # Remove the next run from the paragraph
        paragraph.remove(nxt)
        children = list(paragraph)
        # Don't increment i — check if the next element can also be merged
    return paragraph


def merge_all_runs(tree):
    """Merge fragmented runs in every paragraph of the document."""
    root = tree.getroot()
    for para in root.iter(f'{W}p'):
        merge_runs(para)
    return tree


def tracked_replace_across_runs(paragraph, old_text, new_text):
    """
    Replace `old_text` with `new_text` across potentially multiple runs in a
    paragraph, using tracked changes. Handles the case where Word splits text
    across multiple <w:r> elements.

    Returns True if a replacement was made.
    """
    full_text = get_paragraph_text(paragraph)
    if old_text not in full_text:
        return False

    # First try single-run replacement
    runs = paragraph.findall(f'{W}r')
    for run in runs:
        if tracked_replace_in_run(run, old_text, new_text, paragraph):
            return True

    # Multi-run replacement: find which runs contain pieces of old_text
    run_data = []  # [(run_element, text, start_offset_in_paragraph)]
    offset = 0
    for child in list(paragraph):
        if child.tag == f'{W}r':
            t_elem = child.find(f'{W}t')
            text = t_elem.text if t_elem is not None and t_elem.text else ''
            run_data.append((child, text, offset))
            offset += len(text)
        # Skip non-run elements (they don't contribute to text matching here)

    match_start = full_text.index(old_text)
    match_end = match_start + len(old_text)

    # Find runs that overlap with the match
    affected_runs = []
    for run, text, start in run_data:
        run_end = start + len(text)
        if run_end > match_start and start < match_end:
            affected_runs.append((run, text, start))

    if not affected_runs:
        return False

    # Get formatting from first affected run
    rpr = affected_runs[0][0].find(f'{W}rPr')

    # Build the replacement
    first_run = affected_runs[0]
    insert_anchor = first_run[0]
    anchor_index = list(paragraph).index(insert_anchor)

    # Remove all affected runs
    for run, _, _ in affected_runs:
        paragraph.remove(run)

    insert_pos = anchor_index

    # Text before the match in the first affected run
    first_run_text = first_run[1]
    first_run_start = first_run[2]
    before_text = first_run_text[:match_start - first_run_start]
    if before_text:
        paragraph.insert(insert_pos, make_run_with_rpr(before_text, rpr))
        insert_pos += 1

    # Tracked delete of old text
    r_old = make_run_with_rpr(old_text, rpr)
    paragraph.insert(insert_pos, make_tracked_delete(r_old))
    insert_pos += 1

    # Tracked insert of new text
    r_new = make_run_with_rpr(new_text, rpr)
    paragraph.insert(insert_pos, make_tracked_insert(r_new))
    insert_pos += 1

    # Text after the match in the last affected run
    last_run = affected_runs[-1]
    last_run_text = last_run[1]
    last_run_start = last_run[2]
    last_run_end = last_run_start + len(last_run_text)
    after_text = last_run_text[match_end - last_run_start:] if match_end < last_run_end else ''
    if after_text:
        paragraph.insert(insert_pos, make_run_with_rpr(after_text, rpr))

    return True


def tracked_delete_paragraph(paragraph, body):
    """
    Mark an entire paragraph as deleted using tracked changes.
    Wraps all runs in <w:del>, removes non-run inline elements
    (bookmarkStart, bookmarkEnd, proofErr) that cannot live inside w:del,
    and adds <w:pPr><w:rPr><w:del/></w:rPr></w:pPr> to mark the paragraph
    mark as deleted.
    """
    # Elements that must be removed (cannot be inside w:del, and are not
    # meaningful once the paragraph is deleted)
    remove_tags = {
        f'{W}bookmarkStart', f'{W}bookmarkEnd',
        f'{W}proofErr',
        f'{W}permStart', f'{W}permEnd',
    }

    children = list(paragraph)
    for child in children:
        if child.tag == f'{W}pPr':
            continue  # Keep paragraph properties
        if child.tag in remove_tags:
            paragraph.remove(child)
        elif child.tag == f'{W}r':
            idx = list(paragraph).index(child)
            paragraph.remove(child)
            del_elem = make_tracked_delete(child)
            paragraph.insert(idx, del_elem)
        else:
            # Any other unexpected element — remove it to avoid corruption
            paragraph.remove(child)

    # Mark the paragraph itself as deleted (paragraph mark)
    ppr = paragraph.find(f'{W}pPr')
    if ppr is None:
        ppr = etree.SubElement(paragraph, f'{W}pPr')
        paragraph.insert(0, ppr)

    rpr = ppr.find(f'{W}rPr')
    if rpr is None:
        rpr = etree.SubElement(ppr, f'{W}rPr')

    del_mark = etree.SubElement(rpr, f'{W}del')
    del_mark.set(f'{W}id', next_change_id())
    del_mark.set(f'{W}author', get_author())
    del_mark.set(f'{W}date', get_date())


# ── Structured Markdown data helpers ─────────────────────────────────────────

def _parse_value(v):
    """Parse a string value into appropriate Python type."""
    if v.lower() == 'true':
        return True
    if v.lower() == 'false':
        return False
    if v == '':
        return ''
    return v


def load_md(filepath):
    """Load a structured Markdown data file and return its contents as a dict.

    Format rules:
        # Title          → ignored
        > comment        → ignored
        key: value       → key-value pair in current context
        ## section       → starts a dict section
        ### item         → starts a list-of-dicts item within current section
        key:             → bare key, starts a list (next lines are - items)
        - item           → list item (string)
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    result = {}
    context = result          # dict we're currently adding key-values to
    current_section = None    # name of current ## section (or None for top-level)
    current_list_key = None   # bare "key:" waiting for - items
    current_h3 = None         # dict being built under a ### heading
    h3_target = None          # the list that h3 dicts get appended to

    def _flush_h3():
        nonlocal current_h3, h3_target
        if current_h3 is not None and h3_target is not None:
            h3_target.append(current_h3)
            current_h3 = None

    def _flush_list_key():
        """If a bare key: was set but got no - items, convert [] to ''."""
        nonlocal current_list_key
        if current_list_key and isinstance(context.get(current_list_key), list):
            if len(context[current_list_key]) == 0:
                context[current_list_key] = ''
        current_list_key = None

    for raw_line in lines:
        stripped = raw_line.strip()

        # Skip empty lines and comments
        if not stripped:
            continue
        if stripped.startswith('> '):
            continue

        # h1 title — skip
        if stripped.startswith('# ') and not stripped.startswith('## '):
            continue

        # h2 section
        if stripped.startswith('## ') and not stripped.startswith('### '):
            _flush_h3()
            if current_list_key:
                _flush_list_key()
            section_name = stripped[3:].strip()
            current_section = section_name
            result[section_name] = {}
            context = result[section_name]
            current_list_key = None
            current_h3 = None
            h3_target = None
            continue

        # h3 list-of-dicts item
        if stripped.startswith('### '):
            _flush_h3()
            if current_list_key:
                _flush_list_key()
            # Convert section from dict to list if this is the first h3
            if current_section and isinstance(result.get(current_section), dict) and not result[current_section]:
                result[current_section] = []
                h3_target = result[current_section]
            elif current_section and isinstance(result.get(current_section), list):
                h3_target = result[current_section]
            current_h3 = {}
            context = current_h3
            current_list_key = None
            continue

        # List item
        if stripped.startswith('- '):
            item_text = stripped[2:].strip()
            if current_list_key:
                target = context
                if isinstance(target, dict) and current_list_key in target:
                    target[current_list_key].append(item_text)
            continue

        # Key-value pair (contains ": " with space after colon)
        colon_pos = stripped.find(': ')
        if colon_pos > 0:
            if current_list_key:
                _flush_list_key()
            k = stripped[:colon_pos].strip()
            v = stripped[colon_pos + 2:].strip()
            context[k] = _parse_value(v)
            continue

        # Bare key ending with colon (starts a list)
        if stripped.endswith(':') and not stripped.startswith('#'):
            if current_list_key:
                _flush_list_key()
            key_name = stripped[:-1].strip()
            current_list_key = key_name
            context[key_name] = []
            continue

    # Flush any remaining state
    _flush_h3()
    if current_list_key:
        _flush_list_key()

    return result


def _format_value(v):
    """Format a Python value for Markdown output."""
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if v is None:
        return ''
    return str(v)


def _write_dict_to_lines(lines, d):
    """Write a flat dict's contents as key-value lines and list blocks."""
    for k, v in d.items():
        if isinstance(v, list):
            lines.append('')
            lines.append(f'{k}:')
            for item in v:
                lines.append(f'- {item}')
        else:
            lines.append(f'{k}: {_format_value(v)}')


def save_md(data, filepath, title=None):
    """Save a dict to a structured Markdown data file.

    Handles: top-level scalars, dict sections (## heading),
    list-of-dicts sections (## heading + ### items),
    and list-of-strings within dicts (key: + - items).
    """
    lines = []
    if title:
        lines.append(f'# {title}')
        lines.append('')

    # Top-level scalars first
    for k, v in data.items():
        if not isinstance(v, (dict, list)):
            lines.append(f'{k}: {_format_value(v)}')

    # Then sections (dict and list values)
    for k, v in data.items():
        if isinstance(v, dict):
            lines.append('')
            lines.append(f'## {k}')
            lines.append('')
            _write_dict_to_lines(lines, v)
        elif isinstance(v, list):
            lines.append('')
            lines.append(f'## {k}')
            lines.append('')
            if v and isinstance(v[0], dict):
                for i, item in enumerate(v, 1):
                    lines.append(f'### {i}')
                    lines.append('')
                    _write_dict_to_lines(lines, item)
                    lines.append('')
            else:
                for item in v:
                    lines.append(f'- {item}')

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines) + '\n')


# ── File discovery ───────────────────────────────────────────────────────────

def find_files(folder, extensions=None):
    """
    Recursively find files in a folder.
    If extensions is provided (e.g. ['.pdf', '.docx']), filter by those.
    Returns a list of absolute paths.
    """
    if extensions:
        extensions = [e.lower() for e in extensions]
    results = []
    for root, dirs, files in os.walk(folder):
        for f in files:
            if f.startswith('~$'):
                continue  # Skip Office temp files
            if extensions is None or os.path.splitext(f)[1].lower() in extensions:
                results.append(os.path.join(root, f))
    return sorted(results)


# ── docx text extraction (without python-docx to avoid dependency issues) ──

def extract_docx_text(filepath):
    """Extract plain text from a .docx file using ZIP + XML parsing."""
    import zipfile
    with zipfile.ZipFile(filepath) as z:
        with z.open('word/document.xml') as f:
            tree = etree.parse(f)
    root = tree.getroot()
    texts = []
    for t in root.iter(f'{W}t'):
        if t.text:
            texts.append(t.text)
    return ' '.join(texts)


# ── xlsx text extraction ────────────────────────────────────────────────────

def extract_xlsx_text(filepath):
    """Extract all cell values from an xlsx file as a flat string."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    texts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    texts.append(str(cell))
    wb.close()
    return ' '.join(texts)


# ── PDF splitting for large files ────────────────────────────────────────────

def split_pdf(pdf_path, pages_per_chunk=5):
    """
    Split a PDF into smaller chunk files for reading large documents.

    Args:
        pdf_path: Absolute path to the PDF file.
        pages_per_chunk: Number of pages per chunk file (default 5).

    Returns:
        A tuple of (chunk_paths, chunk_dir) where chunk_paths is a list of
        absolute paths to the chunk PDFs and chunk_dir is the temporary
        directory containing them (pass to cleanup_pdf_chunks when done).
    """
    from pypdf import PdfReader, PdfWriter

    reader = PdfReader(pdf_path)
    total_pages = len(reader.pages)

    if total_pages == 0:
        return [], None

    # Create chunk directory next to the original file
    parent_dir = os.path.dirname(pdf_path)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    chunk_dir = os.path.join(parent_dir, f'_pdf_chunks_{base_name}')
    os.makedirs(chunk_dir, exist_ok=True)

    chunk_paths = []
    for start in range(0, total_pages, pages_per_chunk):
        end = min(start + pages_per_chunk, total_pages)
        writer = PdfWriter()
        for page_num in range(start, end):
            writer.add_page(reader.pages[page_num])

        chunk_filename = f'{base_name}_p{start + 1}-{end}.pdf'
        chunk_path = os.path.join(chunk_dir, chunk_filename)
        with open(chunk_path, 'wb') as f:
            writer.write(f)
        chunk_paths.append(chunk_path)

    return chunk_paths, chunk_dir


def cleanup_pdf_chunks(chunk_dir):
    """Remove a temporary chunk directory created by split_pdf()."""
    if chunk_dir and os.path.isdir(chunk_dir):
        shutil.rmtree(chunk_dir)


# ── Word table helpers ────────────────────────────────────────────────────────

def find_table_by_text(root, search_text):
    """Find a <w:tbl> element containing search_text. Returns the element or None."""
    for tbl in root.iter(f'{W}tbl'):
        for t in tbl.iter(f'{W}t'):
            if t.text and search_text in t.text:
                return tbl
    return None


def get_table_rows(table):
    """Return list of <w:tr> elements in a table."""
    return table.findall(f'{W}tr')


def get_row_cells(row):
    """Return list of <w:tc> elements in a table row."""
    return row.findall(f'{W}tc')


def get_cell_text(cell):
    """Extract all text from a <w:tc> element."""
    texts = []
    for t in cell.iter(f'{W}t'):
        if t.text:
            texts.append(t.text)
    return ''.join(texts)


def get_row_text(row):
    """Extract all text from a table row, returned as list of cell strings."""
    return [get_cell_text(tc) for tc in get_row_cells(row)]


def set_cell_text(cell, text):
    """Replace all text in a <w:tc> element with new text. Preserves first paragraph's formatting."""
    paras = cell.findall(f'{W}p')
    if not paras:
        p = etree.SubElement(cell, f'{W}p')
        r = etree.SubElement(p, f'{W}r')
        t = etree.SubElement(r, f'{W}t')
        t.text = text
        if text and (text[0] == ' ' or text[-1] == ' '):
            t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
        return

    # Use first paragraph, preserve its pPr and first run's rPr
    first_para = paras[0]
    ppr = first_para.find(f'{W}pPr')

    # Grab rPr from first run if available
    first_run = first_para.find(f'{W}r')
    rpr = first_run.find(f'{W}rPr') if first_run is not None else None

    # Clear all children except pPr
    for child in list(first_para):
        if child.tag != f'{W}pPr':
            first_para.remove(child)

    # Add new run with text
    r = etree.SubElement(first_para, f'{W}r')
    if rpr is not None:
        r.append(copy.deepcopy(rpr))
    t = etree.SubElement(r, f'{W}t')
    t.text = text
    if text and (text[0] == ' ' or text[-1] == ' '):
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    # Remove extra paragraphs
    for p in paras[1:]:
        cell.remove(p)


def find_row_by_text(table, search_text):
    """Find first row containing search_text. Returns (index, row_element) or (None, None)."""
    for idx, row in enumerate(get_table_rows(table)):
        row_str = ''.join(get_row_text(row))
        if search_text in row_str:
            return idx, row
    return None, None


def remove_table_rows(table, start_idx, end_idx=None):
    """Remove rows from start_idx to end_idx (exclusive). If end_idx is None, removes to end."""
    rows = get_table_rows(table)
    if end_idx is None:
        end_idx = len(rows)
    for row in rows[start_idx:end_idx]:
        table.remove(row)
    return end_idx - start_idx


def add_table_row(table, cell_values, after_idx=None, clone_format_from=None):
    """
    Add a new row with the given cell text values.
    If clone_format_from is a <w:tr> element, clone its structure for formatting.
    Inserts after after_idx, or appends if None.
    """
    if clone_format_from is not None:
        new_tr = copy.deepcopy(clone_format_from)
        cells = get_row_cells(new_tr)
        for i, val in enumerate(cell_values):
            if i < len(cells):
                set_cell_text(cells[i], str(val))
    else:
        new_tr = etree.Element(f'{W}tr')
        for val in cell_values:
            tc = etree.SubElement(new_tr, f'{W}tc')
            p = etree.SubElement(tc, f'{W}p')
            r = etree.SubElement(p, f'{W}r')
            t = etree.SubElement(r, f'{W}t')
            t.text = str(val)
            if val and (str(val)[0] == ' ' or str(val)[-1] == ' '):
                t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    if after_idx is not None:
        rows = get_table_rows(table)
        if after_idx < len(rows):
            ref_row = rows[after_idx]
            ref_row.addnext(new_tr)
        else:
            table.append(new_tr)
    else:
        table.append(new_tr)
    return new_tr


def remove_rows_by_text(table, search_text):
    """Remove all rows containing search_text. Returns count removed."""
    removed = 0
    for row in list(get_table_rows(table)):
        row_str = ''.join(get_row_text(row))
        if search_text in row_str:
            table.remove(row)
            removed += 1
    return removed


# ── Paragraph section helpers ────────────────────────────────────────────────

def find_paragraphs_between(body, start_text, end_text, inclusive_start=True, inclusive_end=False):
    """
    Find all <w:p> elements between paragraphs containing start_text and end_text.

    Args:
        body: The <w:body> element to search
        start_text: Text to match in the starting paragraph
        end_text: Text to match in the ending paragraph
        inclusive_start: Include the start paragraph in results (default True)
        inclusive_end: Include the end paragraph in results (default False)

    Returns:
        List of <w:p> elements in the range, or empty list if not found.
    """
    paragraphs = list(body.iter(f'{W}p'))
    start_idx = None
    end_idx = None

    for i, para in enumerate(paragraphs):
        text = get_paragraph_text(para)
        if start_idx is None and start_text in text:
            start_idx = i
        elif start_idx is not None and end_text in text:
            end_idx = i
            break

    if start_idx is None:
        return []

    if end_idx is None:
        end_idx = len(paragraphs)

    actual_start = start_idx if inclusive_start else start_idx + 1
    actual_end = end_idx + 1 if inclusive_end else end_idx

    return paragraphs[actual_start:actual_end]


def remove_paragraphs(body, paragraphs):
    """
    Remove a list of <w:p> elements from the body.
    Handles paragraphs inside table cells by replacing with empty <w:p>.

    Args:
        body: The <w:body> element
        paragraphs: List of <w:p> elements to remove

    Returns:
        Count of paragraphs removed.
    """
    TC_TAG = f'{W}tc'
    removed = 0
    for para in paragraphs:
        parent = para.getparent()
        if parent is None:
            continue
        if parent.tag == TC_TAG:
            # Inside a table cell — replace with empty paragraph
            idx = list(parent).index(para)
            parent.remove(para)
            empty_p = etree.Element(f'{W}p')
            parent.insert(idx, empty_p)
        else:
            parent.remove(para)
        removed += 1
    return removed


def make_paragraph(text, bold=False, font_size=None, clone_format_from=None):
    """
    Create a new <w:p> element with the given text.

    Args:
        text: The text content
        bold: Whether to make the text bold
        font_size: Font size in half-points (e.g., 24 = 12pt)
        clone_format_from: A <w:p> element to clone pPr and rPr from

    Returns:
        A new <w:p> element.
    """
    p = etree.Element(f'{W}p')

    if clone_format_from is not None:
        # Clone paragraph properties
        src_ppr = clone_format_from.find(f'{W}pPr')
        if src_ppr is not None:
            p.append(copy.deepcopy(src_ppr))
        # Get run properties from source
        src_run = clone_format_from.find(f'{W}r')
        src_rpr = src_run.find(f'{W}rPr') if src_run is not None else None
    else:
        src_rpr = None

    r = etree.SubElement(p, f'{W}r')

    if src_rpr is not None:
        r.append(copy.deepcopy(src_rpr))
    elif bold or font_size:
        rpr = etree.SubElement(r, f'{W}rPr')
        if bold:
            etree.SubElement(rpr, f'{W}b')
        if font_size:
            sz = etree.SubElement(rpr, f'{W}sz')
            sz.set(f'{W}val', str(font_size))

    t = etree.SubElement(r, f'{W}t')
    t.text = text
    if text and (text[0] == ' ' or text[-1] == ' '):
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    return p


def insert_paragraphs_after(body, anchor_para, new_paragraphs):
    """
    Insert a list of <w:p> elements after a given anchor paragraph.

    Args:
        body: The <w:body> element
        anchor_para: The <w:p> element to insert after
        new_paragraphs: List of <w:p> elements to insert

    Returns:
        Count of paragraphs inserted.
    """
    for i, para in enumerate(new_paragraphs):
        anchor_para.addnext(para)
        anchor_para = para
    return len(new_paragraphs)
